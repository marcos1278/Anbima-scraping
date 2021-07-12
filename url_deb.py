# -*- coding: utf-8 -*-
"""
Created on Sat Jul  3 13:11:01 2021

@author: Marcos
"""

import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import openpyxl
from openpyxl import load_workbook
import pandas as pd

chromedriver = r'C:/Users/Marcos/Downloads/chromedriver'
driver = webdriver.Chrome(executable_path=chromedriver)

wb = load_workbook('D:/2020/Finance/Projeto FICC/Debentures/nome_deb.xlsx')
lenght = (wb.worksheets[0].max_row)
print(lenght)

lista = []

def get_url_debentures(name):
    driver.get("https://data.anbima.com.br/debentures") #Abre a página
    driver.maximize_window() #Maximiza a janela
    
    input_element = driver.find_element_by_class_name("anbima-ui-input-search__text") #Encontra a barra de pesquisa
    input_element.send_keys(name) #Digita na barra de pesquisa
    input_element.send_keys(Keys.RETURN) #Tecla ENTER
    
    # Mudo o driver.get para a página atual de pesquisa que foi aberta
    pagina_atual = driver.window_handles[0]
    driver.switch_to.window(pagina_atual)
    
    driver.implicitly_wait(10) #Espera carregar a página
    
    link = driver.find_element_by_xpath('//*[@id="item-title-0"]') #Encontra o link da página de características
    link.click() #Acessa o link
    
    url = driver.current_url #Pega a url atual
    
    time.sleep(2)
    driver.get('https://data.anbima.com.br/debentures') #Retornar pra página inicial

    return url

for i in range(0, lenght):
    card = {}
    try:
        name = wb.worksheets[0].cell(row=i+1, column=1).value
        card['nome'] = name
        card['url'] = get_url_debentures(name)
        lista.append(card)
            
        #print(get_url_debentures(name))
    except:
        print("erro: " + name)
        continue
    

df = pd.DataFrame(lista)

df.to_excel('deb_url.xlsx')
